#!/usr/bin/env python3
"""
Расчёт порядка драфта свободных агентов по data/draft_config.json.

  S = weight(tier) + weight(league) + (10 - place) + weight(cl)

Меньший S → более ранний пик (1, 2, …). Равенство: меньше очков в лиге,
затем хуже разница мячей.

Запуск: python draft_rating.py [--json-out PATH] [--xlsx-out PATH]
       python draft_rating.py --from-json data/draft_order_result.json --xlsx-out data/draft_order.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent
DEFAULT_CONFIG = ROOT / "data" / "draft_config.json"


def write_draft_xlsx(path: Path, payload: dict) -> None:
    """Таблица Excel: лист «Раунд 1», лист «Раунд 2»."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise SystemExit(
            "Нужен пакет openpyxl: pip install openpyxl"
        ) from e

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Раунд 1"
    ws1.append(["Пик", "Команда", "Лига", "Тир", "Место", "ЛЧ", "Очки", "РМ", "S"])
    for row in payload["round1"]:
        ws1.append(
            [
                row["pick"],
                row["name"],
                row["league"],
                row["tier"],
                row["place"],
                row["cl"],
                row["pts"],
                row["gd"],
                row["S"],
            ]
        )

    ws2 = wb.create_sheet("Раунд 2")
    ws2.append(["Пик", "Команда"])
    for row in payload["round2"]:
        ws2.append([row["pick"], row["name"]])

    for ws in (ws1, ws2):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=False
                )
        for idx in range(1, ws.max_column + 1):
            letter = get_column_letter(idx)
            max_len = len(str(ws.cell(row=1, column=idx).value or ""))
            for r in range(2, ws.max_row + 1):
                max_len = max(max_len, len(str(ws.cell(row=r, column=idx).value or "")))
            ws.column_dimensions[letter].width = min(max_len + 2, 42)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def load_config(path: Path) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def compute_rows(cfg: dict) -> list[dict]:
    w = cfg["weights"]
    tier_w = w["tier"]
    league_w = w["league"]
    cl_w = w["cl"]

    rows = []
    for t in cfg["teams"]:
        if str(t.get("name", "")).startswith("_"):
            continue
        tier = t["tier"]
        lg = t["league"]
        place = int(t["place"])
        cl = t.get("cl") or "none"

        score = (
            tier_w[tier]
            + league_w[lg]
            + (10 - place)
            + cl_w[cl]
        )
        rows.append(
            {
                "name": t["name"],
                "league": lg,
                "tier": tier,
                "place": place,
                "cl": cl,
                "pts": int(t["pts"]) if t.get("pts") is not None else 999,
                "gd": int(t["gd"]) if t.get("gd") is not None else 999,
                "S": round(score, 6),
            }
        )
    return rows


def sort_key(r: dict) -> tuple:
    """Ранний пик: минимальный S; при равенстве — хуже чемпионат (меньше очков, хуже РМ)."""
    return (r["S"], r["pts"], r["gd"])


def main() -> None:
    ap = argparse.ArgumentParser(description="Порядок драфта по draft_config.json")
    ap.add_argument(
        "--config",
        type=Path,
        default=DEFAULT_CONFIG,
        help="Путь к JSON конфигу",
    )
    ap.add_argument(
        "--from-json",
        type=Path,
        default=None,
        help="Только сконвертировать готовый draft_order JSON в Excel (без пересчёта)",
    )
    ap.add_argument("--json-out", type=Path, default=None, help="Сохранить результат в JSON")
    ap.add_argument(
        "--xlsx-out",
        type=Path,
        default=None,
        help="Сохранить результат в Excel (.xlsx)",
    )
    args = ap.parse_args()

    if args.from_json is not None:
        with open(args.from_json, encoding="utf-8") as f:
            payload = json.load(f)
        if args.xlsx_out is None:
            raise SystemExit("С --from-json укажите --xlsx-out")
        write_draft_xlsx(args.xlsx_out, payload)
        print(f"[записано {args.xlsx_out}]", file=sys.stderr)
        return

    cfg = load_config(args.config)
    rows = compute_rows(cfg)
    if len(rows) != 50:
        raise SystemExit(f"Ожидается 50 команд, в конфиге: {len(rows)}")

    rows.sort(key=sort_key)

    lines = []
    lines.append("=== Раунд 1 (пики 1–50) ===\n")
    for i, r in enumerate(rows, start=1):
        lines.append(
            f"{i:2d} — {r['name']:<16}  S={r['S']:.0f}  "
            f"(тир {r['tier']}, лига {r['league']}, место {r['place']}, ЛЧ {r['cl']})"
        )

    lines.append("\n=== Раунд 2 (51–100), snake ===\n")
    order_r1 = [r["name"] for r in rows]
    for k in range(50):
        pick = 51 + k
        name = order_r1[49 - k]
        lines.append(f"{pick} — {name}")

    text = "\n".join(lines)
    print(text)

    out = None
    if args.json_out or args.xlsx_out:
        out = {
            "round1": [{"pick": i + 1, **rows[i]} for i in range(50)],
            "round2": [
                {"pick": 51 + k, "name": order_r1[49 - k]} for k in range(50)
            ],
            "formula": "S = tier + league + (10 - place) + cl; sort ascending S, then pts, then gd",
        }
    if args.json_out and out is not None:
        args.json_out.parent.mkdir(parents=True, exist_ok=True)
        with open(args.json_out, "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False, indent=2)
        print(f"\n[записано {args.json_out}]")
    if args.xlsx_out and out is not None:
        write_draft_xlsx(args.xlsx_out, out)
        print(f"\n[записано {args.xlsx_out}]")


if __name__ == "__main__":
    main()
