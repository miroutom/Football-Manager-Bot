#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel-таблица составов 40 клубов для ручного трансферного окна.

  python3 scripts/export_transfer_window_squads_xlsx.py
  python3 scripts/export_transfer_window_squads_xlsx.py -o data/transfer_window_squads.xlsx

Лист «Окно»: команды сеткой 8×5, блок = заголовок + старт (11) + запас (7) + резерв (до 9).
Счётчики IN/OUT и зелёная подсветка пришедших — по сравнению с листом «_orig» (исходные составы).
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from player_stats import LEAGUE_NAMES, LEAGUE_TEAMS
from utils.roster_manual import _iter_team_players
from utils.season_paths import get_active_season
from utils.transfer_market_draft import _EXCLUDED_TEAMS
from utils.utils import session_league

N_START = 11
N_BENCH = 7
N_RESERVE = 9
N_PLAYER_SLOTS = N_START + N_BENCH + N_RESERVE
COLS_PER_TEAM = 3
GAP_COLS = 2
TEAM_GRID_W = COLS_PER_TEAM + GAP_COLS
BLOCK_H = 1 + 1 + N_START + 1 + N_BENCH + 1 + N_RESERVE  # header+labels+slots
GAP_ROWS = 2
ROW_STEP = BLOCK_H + GAP_ROWS
TEAMS_PER_ROW = 8

_FONT_HDR = Font(bold=True, size=9)
_FONT_LABEL = Font(bold=True, size=8, color="444444")
_FONT_DATA = Font(size=8)
_FILL_HDR = PatternFill("solid", fgColor="D9E1F2")
_FILL_LABEL = PatternFill("solid", fgColor="F2F2F2")
_FILL_IN = PatternFill("solid", fgColor="C6EFCE")
_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _teams_40() -> list[str]:
    out: list[str] = []
    for code in ("rpl", "eng", "esp", "ita", "ger"):
        for team in LEAGUE_TEAMS.get(code, []):
            if team not in _EXCLUDED_TEAMS:
                out.append(team)
    return out


def _load_squad(team: str) -> tuple[list[tuple], list[tuple], list[tuple]]:
    buckets: dict[str, list[tuple[str, str, int]]] = {
        "start": [],
        "bench": [],
        "reserve": [],
    }
    for _, r in _iter_team_players(session_league, team):
        if getattr(r, "left_team", False):
            continue
        st = (getattr(r, "status", None) or "reserve").strip().lower()
        if st not in buckets:
            st = "reserve"
        buckets[st].append((r.name or "", r.position or "", int(r.overall or 0)))
    for key in buckets:
        buckets[key].sort(key=lambda x: (-x[2], x[0]))
    return buckets["start"], buckets["bench"], buckets["reserve"]


def _team_origin_col(team_idx: int) -> int:
    return team_idx + 1


def _write_orig_sheet(ws, teams: list[str], squads: dict[str, tuple]) -> dict[str, str]:
    """Лист _orig: по колонке на клуб. Возвращает col_letter для каждой команды."""
    ws.sheet_state = "hidden"
    letters: dict[str, str] = {}
    for i, team in enumerate(teams):
        col = _team_origin_col(i)
        letter = get_column_letter(col)
        letters[team] = letter
        ws.cell(1, col, team)
        ws.cell(1, col).font = _FONT_HDR
        start, bench, reserve = squads[team]
        names = [p[0] for p in start + bench + reserve if p[0]]
        for row_off, name in enumerate(names, start=2):
            ws.cell(row_off, col, name)
        ws.column_dimensions[letter].width = 16
    return letters


def _player_name_range(letter: str, first_row: int, last_row: int) -> str:
    return f"{letter}${first_row}:{letter}${last_row}"


def _orig_range(letter: str, n_names: int) -> str:
    last = max(2, 1 + n_names)
    return f"_orig!${letter}$2:${letter}${last}"


def _counter_formula(team: str, name_rng: str, orig_rng: str) -> str:
    in_part = (
        f'MIN(5,SUMPRODUCT((LEN({name_rng})>0)*'
        f'(ISERROR(MATCH({name_rng},{orig_rng},0)))))'
    )
    out_part = (
        f'MIN(5,SUMPRODUCT((LEN({orig_rng})>0)*'
        f'(ISERROR(MATCH({orig_rng},{name_rng},0)))))'
    )
    return f'="{team} "&{in_part}&"/5 IN "&{out_part}&"/5 OUT"'


def _place_team(
    ws,
    team: str,
    start_row: int,
    start_col: int,
    start: list[tuple],
    bench: list[tuple],
    reserve: list[tuple],
    orig_letter: str,
) -> None:
    name_col = start_col
    pos_col = start_col + 1
    ovr_col = start_col + 2
    name_letter = get_column_letter(name_col)

    # Имена только в слотах игроков; подписи секций — в колонке позиции (имя пустое).
    first_player_row = start_row + 2
    last_player_row = start_row + BLOCK_H - 1

    all_names = [p[0] for p in start + bench + reserve if p[0]]
    name_rng = _player_name_range(name_letter, first_player_row, last_player_row)
    orig_rng = _orig_range(orig_letter, len(all_names))

    hdr = ws.cell(start_row, name_col, None)
    hdr.value = _counter_formula(team, name_rng, orig_rng)
    hdr.font = _FONT_HDR
    hdr.fill = _FILL_HDR
    ws.merge_cells(
        start_row=start_row,
        start_column=name_col,
        end_row=start_row,
        end_column=ovr_col,
    )

    def _label(row: int, text: str) -> None:
        c = ws.cell(row, pos_col, text)
        c.font = _FONT_LABEL
        c.fill = _FILL_LABEL
        ws.merge_cells(
            start_row=row, start_column=pos_col, end_row=row, end_column=ovr_col
        )

    def _players(row: int, players: list[tuple], slots: int) -> int:
        r = row
        for i in range(slots):
            if i < len(players):
                nm, pos, ovr = players[i]
            else:
                nm, pos, ovr = "", "", ""
            for col, val in ((name_col, nm), (pos_col, pos), (ovr_col, ovr)):
                cell = ws.cell(r, col, val if val != "" else None)
                cell.font = _FONT_DATA
                cell.alignment = _ALIGN
            r += 1
        return r

    r = start_row + 1
    _label(r, "старт")
    r = _players(r + 1, start, N_START)
    _label(r, "запас")
    r = _players(r + 1, bench, N_BENCH)
    _label(r, "резерв")
    _players(r + 1, reserve, N_RESERVE)

    cf_rng = (
        f"{name_letter}{first_player_row}:{name_letter}{last_player_row}"
    )
    rule = FormulaRule(
        formula=[
            f"AND(LEN({name_letter}{first_player_row})>0,"
            f"ISERROR(MATCH({name_letter}{first_player_row},{orig_rng},0)))"
        ],
        fill=_FILL_IN,
    )
    ws.conditional_formatting.add(cf_rng, rule)


def _setup_page(ws, last_col: int, last_row: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.4, bottom=0.4)
    ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"
    for c in range(1, last_col + 1):
        pos_in_block = (c - 1) % TEAM_GRID_W
        if pos_in_block == 0:
            ws.column_dimensions[get_column_letter(c)].width = 14
        elif pos_in_block == 1:
            ws.column_dimensions[get_column_letter(c)].width = 4.5
        elif pos_in_block == 2:
            ws.column_dimensions[get_column_letter(c)].width = 5
        else:
            ws.column_dimensions[get_column_letter(c)].width = 1.2
    for r in range(1, last_row + 1):
        ws.row_dimensions[r].height = 12


def build_xlsx(path: Path) -> None:
    teams = _teams_40()
    if len(teams) != 40:
        raise RuntimeError(f"ожидалось 40 клубов, получено {len(teams)}")

    squads = {t: _load_squad(t) for t in teams}

    wb = Workbook()
    ws_orig = wb.active
    ws_orig.title = "_orig"
    orig_letters = _write_orig_sheet(ws_orig, teams, squads)

    ws = wb.create_sheet("Окно", 0)
    ws["A1"] = (
        "Перетаскивайте строки игроков между клубами. "
        "Зелёный = новый в клубе (IN). Счётчики IN/OUT — до 5."
    )
    ws["A1"].font = Font(bold=True, size=10)
    ws.merge_cells("A1:AN1")

    base_row = 3
    for idx, team in enumerate(teams):
        grid_row = idx // TEAMS_PER_ROW
        grid_col = idx % TEAMS_PER_ROW
        start_row = base_row + grid_row * ROW_STEP
        start_col = 1 + grid_col * TEAM_GRID_W
        start, bench, reserve = squads[team]
        _place_team(
            ws,
            team,
            start_row,
            start_col,
            start,
            bench,
            reserve,
            orig_letters[team],
        )

    last_row = base_row + ((len(teams) - 1) // TEAMS_PER_ROW) * ROW_STEP + BLOCK_H - 1
    last_col = TEAMS_PER_ROW * TEAM_GRID_W
    _setup_page(ws, last_col, last_row)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    ap = argparse.ArgumentParser(description="Excel: составы 40 клубов для трансферного окна")
    ap.add_argument(
        "-o",
        "--output",
        default=str(_ROOT / "data" / "transfer_window_squads.xlsx"),
        help="Путь к .xlsx",
    )
    args = ap.parse_args()
    out = Path(args.output)
    season = get_active_season()
    build_xlsx(out)
    print(f"Сезон {season}: 40 клубов → {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
