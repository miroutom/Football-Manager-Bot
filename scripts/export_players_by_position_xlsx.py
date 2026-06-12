#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Экспорт всех игроков текущего сезона (лига + ЛЧ) в Excel.

  python3 scripts/export_players_by_position_xlsx.py
  python3 scripts/export_players_by_position_xlsx.py -o data/players_s3.xlsx
"""
from __future__ import annotations

import argparse
import os
import sys

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from utils.players_by_position import collect_players_flat
from utils.season_paths import get_active_season

_HEADER = ("Фамилия", "Имя", "Команда", "Позиция", "Рейтинг", "Менеджер")


def _write_xlsx(path: str, rows: list[dict]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise SystemExit("Нужен openpyxl: pip install openpyxl") from e

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Игроки"
    ws.append(list(_HEADER))
    for r in rows:
        ws.append(
            [
                r.get("surname", ""),
                r.get("name", ""),
                r.get("team", ""),
                r.get("position", ""),
                int(r.get("overall", 0) or 0),
                r.get("manager", ""),
            ]
        )
    for col in range(1, len(_HEADER) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.auto_filter.ref = ws.dimensions
    for idx in range(1, len(_HEADER) + 1):
        letter = get_column_letter(idx)
        max_len = len(_HEADER[idx - 1])
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=idx).value
            max_len = max(max_len, len(str(val or "")))
        ws.column_dimensions[letter].width = min(max_len + 2, 40)
    wb.save(path)


def main() -> None:
    ap = argparse.ArgumentParser(description="Excel: игроки по позициям (все в одной таблице)")
    ap.add_argument(
        "-o",
        "--output",
        default="",
        help="Путь к .xlsx (по умолчанию data/players_by_position_season_N.xlsx)",
    )
    args = ap.parse_args()

    season = get_active_season()
    out = args.output.strip() or os.path.join(
        _ROOT, "data", f"players_by_position_season_{season}.xlsx"
    )
    rows = collect_players_flat()
    if not rows:
        print("Нет игроков в БД активного сезона.")
        sys.exit(1)
    _write_xlsx(out, rows)
    print(f"Сезон {season}: {len(rows)} игроков → {out}")


if __name__ == "__main__":
    main()
