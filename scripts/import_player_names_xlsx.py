#!/usr/bin/env python3
"""
Импорт имён из ``db/names.xlsx`` (блоки по командам).

Сезон 2: клуб + позиция + рейтинг + подпись (нация и полное имя в БД — гибко).
Сезон 1: фамилия + нация по всей БД (клуб/позиция/рейтинг — для уточнения).

  python3 scripts/import_player_names_xlsx.py --season 2
  python3 scripts/import_player_names_xlsx.py --season 2 --apply
  python3 scripts/import_player_names_xlsx.py --season 1
"""
from __future__ import annotations

import argparse
import os
import sys
from dataclasses import dataclass

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from data.defender import Defender
from data.forward import Forward
from data.goalkeeper import Goalkeeper
from data.midfielder import Midfielder
from config.squad_team_aliases import canonical_team_name
from utils.migrate_player_surname import prepare_season_archive_schema
from utils.player_names import is_empty_first_name_value
from utils.player_transfer import _norm_cmp, normalize_player_name_for_db

_ALL = (
    (Forward, "forwards"),
    (Midfielder, "midfielders"),
    (Defender, "defenders"),
    (Goalkeeper, "goalkeepers"),
)


@dataclass
class XlsxPlayer:
    team: str
    surname_label: str  # колонка surname в xlsx → целевая фамилия
    position: str
    rating: int
    nation: str
    first_name: str  # колонка name


def _first_name_from_cell(raw: str) -> str:
    s = normalize_player_name_for_db(raw or "")
    if is_empty_first_name_value(s):
        return ""
    return s


# Синонимы национальности (xlsx ↔ БД)
_NATION_CANON: dict[str, str] = {
    "босния и герцеговина": "босния",
    "босния и герцеговна": "босния",
    "босния и герц": "босния",
    "юж. корея": "южная корея",
    "южная корея": "южная корея",
    "кот-д'ивуар": "кот д ивуар",
    "кот-д ивуар": "кот д ивуар",
    "др конго": "др конго",
    "д р конго": "др конго",
    "конго": "др конго",
    "оаэ": "оаэ",
    "сауд. аравия": "саудовская аравия",
    "саудовская аравия": "саудовская аравия",
    "англия": "англия",
    "англ": "англия",
    "корея": "южная корея",
}

# Название клуба в xlsx → как в БД (МЮ и т.п.)
_XLSX_TEAM_CANON: dict[str, str] = {
    "мю": "Мю",
    "манчестер юнайтед": "Мю",
}

# Опечатки в xlsx → подпись в БД (norm_cmp)
_SURNAME_SPELLING_ALIASES: dict[str, str] = {
    "купмайнерс": "копмейнерс",
    "заппакоста": "дзаппакоста",
}

# (клуб, фамилия в xlsx, позиция) → фрагмент имени в БД
_TEAM_POS_DB_HINT: dict[tuple[str, str, str], str] = {
    ("атлетик", "вильямс", "ЛФА"): "нико",
    ("атлетик", "вильямс", "ПФА"): "иньяки",
    ("жирона", "гарсия", "ЦП"): "алейш",
    ("жирона", "гарсия", "ЦЗ"): "эрик",
    # S1: в xlsx блок Вольфсбург, в архиве — Саму в Жироне
    ("вольфсбург", "гарсия", "ЦП"): "саму",
}

# В xlsx есть блок, но в архиве season_1 этого игрока нет (появился в S2)
_SEASON1_XLSX_SKIP: frozenset[tuple[str, str]] = frozenset(
    {
        (_norm_cmp("Аталанта"), _norm_cmp("Сильва")),
    }
)


def _canonical_xlsx_team(team: str) -> str:
    s = (team or "").strip()
    if not s:
        return s
    key = s.casefold()
    if key in _XLSX_TEAM_CANON:
        return _XLSX_TEAM_CANON[key]
    return canonical_team_name(s)


def _norm_nat(s: str) -> str:
    n = (s or "").strip().casefold()
    n = n.replace(".", " ")
    n = " ".join(n.split())
    return _NATION_CANON.get(n, n)


def _db_listing_label(row) -> str:
    """Подпись игрока в БД (поле ``name``)."""
    return (getattr(row, "name", None) or "").strip()


def _db_labels(row) -> set[str]:
    """Все варианты подписи игрока в БД (включая последнее слово составного имени)."""
    fn = (getattr(row, "name", None) or "").strip()
    from utils.player_names import player_surname

    sn = player_surname(row)
    labels = {
        _norm_cmp(_db_listing_label(row)),
        _norm_cmp(fn),
        _norm_cmp(sn),
    }
    for raw in (fn, sn):
        if not raw:
            continue
        parts = raw.split()
        if len(parts) > 1:
            labels.add(_norm_cmp(parts[-1]))
    return {x for x in labels if x}


def _label_matches_db(
    xlsx_surname: str, row, *, entry: XlsxPlayer | None = None
) -> bool:
    """Совпадение подписи в xlsx с тем, что сейчас в БД."""
    want = _norm_cmp(xlsx_surname)
    if not want:
        return True
    labels = _db_labels(row)
    seeks = {want}
    alt = _SURNAME_SPELLING_ALIASES.get(want)
    if alt:
        seeks.add(_norm_cmp(alt))
    for seek in seeks:
        if seek in labels:
            return True
        for lab in labels:
            parts = lab.split()
            if parts and parts[-1] == seek:
                return True
    if entry:
        pos = (entry.position or "").strip().upper()
        hint = _team_pos_hint(entry)
        if hint and any(hint in lab for lab in labels):
            return True
    return False


def _team_pos_hint(entry: XlsxPlayer) -> str | None:
    pos = (entry.position or "").strip().upper()
    want = _norm_cmp(entry.surname_label)
    return _TEAM_POS_DB_HINT.get((_norm_cmp(entry.team), want, pos))


def _narrow_by_team_pos_hint(
    cands: list[tuple[str, object]], entry: XlsxPlayer
) -> list[tuple[str, object]]:
    """Точечные случаи: клуб+фамилия+позиция → один игрок по фрагменту имени в БД."""
    hint = _team_pos_hint(entry)
    if not hint:
        return cands
    filt = [
        c for c in cands if any(hint in lab for lab in _db_labels(c[1]))
    ]
    return filt if filt else cands


def _filter_by_first_name(
    cands: list[tuple[str, object]], entry: XlsxPlayer
) -> list[tuple[str, object]]:
    fn = (entry.first_name or "").strip()
    if not fn:
        return cands
    want = _norm_cmp(fn)
    filt = [
        c
        for c in cands
        if any(
            want in lab or lab.startswith(want)
            for lab in _db_labels(c[1])
        )
    ]
    return filt if filt else cands


def load_names_xlsx(path: str) -> list[XlsxPlayer]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    out: list[XlsxPlayer] = []
    team = ""
    try:
        for row in ws.iter_rows(values_only=True):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            c0 = row[0]
            c1 = row[1] if len(row) > 1 else None
            if c0 is None:
                continue
            s0 = str(c0).strip()
            if c1 is None or str(c1).strip() == "":
                if s0.lower() != "surname":
                    team = _canonical_xlsx_team(s0)
                continue
            if str(c1).strip().lower() == "position":
                continue
            pos = str(c1).strip().upper()
            try:
                rating = int(row[2]) if row[2] is not None else 0
            except (TypeError, ValueError):
                rating = 0
            nation = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
            first_raw = row[4] if len(row) > 4 else ""
            first = _first_name_from_cell(str(first_raw) if first_raw is not None else "")
            sn = normalize_player_name_for_db(s0) or s0
            if not sn or not team:
                continue
            out.append(
                XlsxPlayer(
                    team=team,
                    surname_label=sn,
                    position=pos,
                    rating=rating,
                    nation=nation,
                    first_name=first,
                )
            )
    finally:
        wb.close()
    return out


def _iter_rows(session, team: str | None, *, include_left: bool):
    for Cls, tbl in _ALL:
        for r in session.query(Cls).all():
            if not include_left and bool(getattr(r, "left_team", False)):
                continue
            t = (getattr(r, "team", None) or "").strip()
            if not t or t.casefold() == "free agent":
                continue
            if team is not None and _norm_cmp(t) != _norm_cmp(team):
                continue
            yield tbl, r


def _candidates_surname_nation(session, entry: XlsxPlayer) -> list[tuple[str, object]]:
    out: list[tuple[str, object]] = []
    for tbl, r in _iter_rows(session, None, include_left=False):
        if _norm_nat(getattr(r, "nation", None) or "") != _norm_nat(entry.nation):
            continue
        if not _label_matches_db(entry.surname_label, r, entry=entry):
            continue
        out.append((tbl, r))
    return out


def _candidates_team_label(
    session, entry: XlsxPlayer, *, require_nation: bool
) -> list[tuple[str, object]]:
    out: list[tuple[str, object]] = []
    for tbl, r in _iter_rows(session, entry.team, include_left=False):
        if require_nation and _norm_nat(
            getattr(r, "nation", None) or ""
        ) != _norm_nat(entry.nation):
            continue
        if not _label_matches_db(entry.surname_label, r, entry=entry):
            continue
        out.append((tbl, r))
    return _filter_by_first_name(out, entry)


def _candidates_team_pos_rating_label(
    session, entry: XlsxPlayer, *, require_nation: bool
) -> list[tuple[str, object]]:
    """Клуб + позиция + рейтинг + подпись (нация опционально)."""
    out: list[tuple[str, object]] = []
    for tbl, r in _iter_rows(session, entry.team, include_left=False):
        if (getattr(r, "position", None) or "").strip().upper() != entry.position:
            continue
        if int(getattr(r, "overall", 0) or 0) != entry.rating:
            continue
        if require_nation and _norm_nat(
            getattr(r, "nation", None) or ""
        ) != _norm_nat(entry.nation):
            continue
        if not _label_matches_db(entry.surname_label, r, entry=entry):
            continue
        out.append((tbl, r))
    return _filter_by_first_name(out, entry)


def _pick_one(
    cands: list[tuple[str, object]], entry: XlsxPlayer, *, prefer_team: bool
) -> tuple[tuple[str, object] | None, str | None]:
    if not cands:
        return None, "не найден"
    if len(cands) == 1:
        return cands[0], None

    def _filt(rows: list[tuple[str, object]], pred) -> list[tuple[str, object]]:
        x = [c for c in rows if pred(c[1])]
        return x if x else rows

    pool = list(cands)
    pool = _filt(
        pool,
        lambda r: (getattr(r, "position", None) or "").strip().upper() == entry.position,
    )
    pool = _filt(pool, lambda r: int(getattr(r, "overall", 0) or 0) == entry.rating)
    if prefer_team:
        pool = _filt(
            pool, lambda r: _norm_cmp(r.team or "") == _norm_cmp(entry.team)
        )
    if len(pool) == 1:
        return pool[0], None
    clubs = sorted({(c[1].team or "").strip() for c in pool if (c[1].team or "").strip()})
    extra = f" ({', '.join(clubs[:4])}{'…' if len(clubs) > 4 else ''})" if clubs else ""
    return None, f"неоднозначно: {len(pool)}{extra}"


def _candidates_team_surname(
    session, entry: XlsxPlayer
) -> list[tuple[str, object]]:
    return _candidates_team_label(session, entry, require_nation=False)


def find_db_row_season1(session, entry: XlsxPlayer):
    """Сезон 1: фамилия+нация по всей БД; клуб из xlsx — для уточнения."""
    if (_norm_cmp(entry.team), _norm_cmp(entry.surname_label)) in _SEASON1_XLSX_SKIP:
        return None, "не найден"

    cands = _candidates_surname_nation(session, entry)
    cands = _narrow_by_team_pos_hint(cands, entry)
    cands = _filter_by_first_name(cands, entry)
    hit, err = _pick_one(cands, entry, prefer_team=True)
    if hit:
        return hit, err
    # Неоднозначно только среди чужих клубов → в S1 в этом клубе игрока не было
    if err and "неоднознач" in err and entry.team and cands:
        at_club = any(
            _norm_cmp(c[1].team or "") == _norm_cmp(entry.team) for c in cands
        )
        if not at_club:
            return None, "не найден"

    if entry.team:
        fb = _candidates_team_label(session, entry, require_nation=False)
        fb = _filter_by_first_name(fb, entry)
        hit, err = _pick_one(fb, entry, prefer_team=False)
        if hit:
            return hit, err

    return None, err or "не найден"


def find_db_row_season2(session, entry: XlsxPlayer):
    """Сезон 2: клуб + позиция + рейтинг + подпись; нация и полное имя — гибко."""
    err: str | None = None
    for require_nat in (True, False):
        cands = _candidates_team_pos_rating_label(
            session, entry, require_nation=require_nat
        )
        if len(cands) == 1:
            return cands[0], None
        if len(cands) > 1:
            hit, err = _pick_one(cands, entry, prefer_team=False)
            if hit:
                return hit, err
            if err and "неоднознач" in err:
                return None, err

    for require_nat in (True, False):
        cands = _candidates_team_label(session, entry, require_nation=require_nat)
        hit, err = _pick_one(cands, entry, prefer_team=False)
        if hit:
            return hit, err
        if err and "неоднознач" in err:
            return None, err

    return None, err or "не найден"


def find_db_row(session, entry: XlsxPlayer, *, require_team: bool):
    if not require_team:
        return find_db_row_season1(session, entry)
    return find_db_row_season2(session, entry)


def apply_names(
    session,
    entries: list[XlsxPlayer],
    *,
    season: int,
    require_team: bool,
    do_apply: bool,
    problems_only: bool = False,
) -> dict[str, int]:
    stats = {"ok": 0, "skip": 0, "miss": 0, "ambig": 0, "team_miss": 0}
    by_team: dict[str, list[str]] = {}
    problems: list[tuple[str, str]] = []

    for e in entries:
        hit, err = find_db_row(session, e, require_team=require_team)
        if hit is None:
            stats["ambig" if err and "неоднознач" in err else "miss"] += 1
            line = f"  ? {e.surname_label} {e.position} {e.rating} {e.nation} — {err}"
            problems.append((e.team, line.strip()))
            by_team.setdefault(e.team, []).append(line)
            continue

        tbl, r = hit
        fn = e.first_name
        sn = e.surname_label
        from utils.player_names import apply_parsed_names_to_row

        old_name = (getattr(r, "name", None) or "").strip()
        new_name = (
            f"{fn.strip().title()} {sn.strip().title()}".strip()
            if fn and sn
            else (sn or fn or "").strip().title()
        )
        if _norm_cmp(old_name) == _norm_cmp(new_name):
            stats["skip"] += 1
            continue

        stats["ok"] += 1
        if problems_only:
            continue
        club = (r.team or e.team).strip()
        line = f"  {tbl} id={r.id} {club}: «{old_name}» → «{new_name}»"
        by_team.setdefault(e.team if require_team else club, []).append(line)
        if do_apply:
            apply_parsed_names_to_row(r, fn, sn)

    if not problems_only:
        for team in sorted(by_team.keys()):
            print(team)
            for line in by_team[team]:
                print(line)
            print()

    if problems:
        print("=" * 60)
        print("НЕ СОПОСТАВЛЕНО (из xlsx → нет строки в БД сезона)")
        print("=" * 60)
        cur = ""
        for team, line in sorted(problems, key=lambda x: (x[0].casefold(), x[1])):
            if team != cur:
                print(f"\n{team}")
                cur = team
            print(f"  {line}")

    return stats


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--xlsx",
        default=os.path.join(ROOT, "db", "names.xlsx"),
        help="Путь к Excel",
    )
    ap.add_argument("--season", type=int, default=2)
    ap.add_argument("--apply", action="store_true")
    ap.add_argument(
        "--problems-only",
        action="store_true",
        help="Только пропуски (?): не найден / неоднозначно",
    )
    args = ap.parse_args()

    if not os.path.isfile(args.xlsx):
        print(f"Нет файла: {args.xlsx}")
        sys.exit(1)

    entries = load_names_xlsx(args.xlsx)
    if not entries:
        print("В xlsx нет строк игроков.")
        sys.exit(1)

    require_team = args.season >= 2
    mode = f"сезон {args.season}" + (
        " + клуб + позиция + рейтинг" if require_team else " — фамилия + нация"
    )
    print(f"{'Запись' if args.apply else 'Просмотр'} ({mode}), строк в файле: {len(entries)}\n")

    prepare_season_archive_schema(args.season)

    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    from utils.utils import Base

    path = os.path.join(ROOT, "db", f"season_{args.season}", "league.db")
    if not os.path.isfile(path):
        print(f"Нет {path}")
        sys.exit(1)

    engine = create_engine(f"sqlite:///{path}")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    try:
        stats = apply_names(
            session,
            entries,
            season=args.season,
            require_team=require_team,
            do_apply=args.apply,
            problems_only=args.problems_only,
        )
        if args.apply:
            session.commit()
            if args.season == 2:
                from utils.common_db import rebuild_common_database

                rebuild_common_database()
                print("common.db пересобран (активный сезон).")
        print(
            f"Итого: обновить {stats['ok']}, без изменений {stats['skip']}, "
            f"не найдено {stats['miss']}, неоднозначно {stats['ambig']}."
        )
        if not args.apply:
            print("\nДля записи добавь --apply")
    finally:
        session.close()
        engine.dispose()


if __name__ == "__main__":
    main()
